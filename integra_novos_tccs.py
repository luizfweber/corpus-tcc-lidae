#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Integra novos TCCs ao corpus consolidado.

USO:
  python3 integra_novos_tccs.py <arquivo_novo.csv>

EXEMPLO:
  python3 integra_novos_tccs.py ../../../Dados\ até\ 12_junho_2026/Catalogação\ de\ TCCs\ –\ 16jun2026v2.csv

O script:
  1. Carrega o novo arquivo CSV (mesmo formato dos formulários originais)
  2. Aplica as mesmas normalizações (limpeza textual, normalização de nomes)
  3. Detecta duplicatas (TCCs que já estão na base)
  4. Adiciona apenas os TCCs novos
  5. Regenera CSV e XLSX consolidados
  6. Cria backup da base anterior
"""
import sys
from pathlib import Path
import pandas as pd
import csv
import re
import unicodedata
from collections import OrderedDict
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# CAMINHOS
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
CONSOLIDADO_CSV = BASE_DIR / "outputs" / "analise" / "corpus_tccs_analisado.csv"
CONSOLIDADO_XLSX = BASE_DIR / "outputs" / "analise" / "corpus_tccs_analisado.xlsx"
BACKUP_DIR = BASE_DIR / "outputs" / "backups"
BACKUP_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# FUNÇÕES (copiadas de consolida_corpus.py)
# ─────────────────────────────────────────────────────────────────────────────
def chave(s):
    """Normaliza para deduplicação."""
    s = (s or "").upper().strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def limpa_texto(s):
    """Remove quebras de linha internas."""
    s = (s or "").replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip()


def limpa_nome(s):
    """Remove títulos e normaliza nomes."""
    s = str(s or "").strip()
    s = s.rstrip(".,;:")
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r",\s*(?:Universidade|Federal|Instituto|UFRR|UFAM|SEED|Curso|Departamento).*$", "", s, flags=re.I)
    s = re.sub(r"\b(Prof|Profa|Professor|Professora|Dr|Dra|Doutor|Doutora|"
                r"Me|Msc|Mestrado|Ms|Mestre|Esp|Ph\.?D|PhD|MSc|Ma|Pós-doutor)\b\.?\s*",
                "", s, flags=re.I)
    s = re.sub(r"\s+(?:em|de Educação|de Educación|de Estudo).*$", "", s, flags=re.I)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace(".", " ")
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = " ".join(w.capitalize() for w in s.split() if w)
    return s


def grupo_canonico(curso):
    """Normaliza nome de curso para grupo canônico."""
    k = chave(curso)
    if k.startswith("PEDAGOGIA"):
        return "Pedagogia"
    if k.startswith("MATEMATICA"):
        return "Matemática"
    if k.startswith("MUSICA"):
        return "Música"
    if k.startswith("INSIKIRAN"):
        return "Insikiran"
    if k.startswith("LEDUCAR"):
        return "LEDUCAR"
    if k.startswith("LETRAS"):
        return "Letras"
    if k.startswith("HISTORIA"):
        return "História"
    if k.startswith("GEOGRAFIA"):
        return "Geografia"
    if k.startswith("CIENCIAS"):
        return "Ciências Biológicas"
    if k.startswith("QUIMICA"):
        return "Química"
    if k.startswith("FISICA"):
        return "Física"
    if k.startswith("ARTES"):
        return "Artes Visuais"
    if k == "OUTRO" or k == "":
        return "Outro (a classificar)"
    return curso.strip()


# ─────────────────────────────────────────────────────────────────────────────
# MAPEAR COLUNAS (como em consolida_corpus.py)
# ─────────────────────────────────────────────────────────────────────────────
MAP = {
    "titulo": "Título completo do TCC",
    "tipo_tcc": "Tipo do TCC",
    "autor": "Autor/a do TCC",
    "orientador": "Orientador/a",
    "coorientador": "Coorientador/a, se houver",
    "banca_examinadora": "Membros da banca examinadora, se houver",
    "ano_defesa": "Ano de defesa",
    "semestre_defesa": "Semestre de defesa, se houver",
    "paginas": "Número de páginas",
    "palavras_chave": "Palavras-chave informadas no TCC",
    "resumo_disponivel": "Resumo disponível?",
    "resumo": "Resumo do TCC",
    "tcc_digital": "TCC digital?",
    "formato": "Formato do documento analisado",
    "disponivel_integral": "TCC disponível integralmente?",
    "link": "Link ou localização do TCC",
    "arquivo": "Envio do arquivo do TCC",
    "observacoes_acesso": "Observações sobre acesso ao documento",
    "observacoes_analise": "Observações da análise",
    "carimbo": "Carimbo de data/hora",
}

TEXTO_LIMPO = {"resumo", "titulo", "palavras_chave", "banca_examinadora",
               "observacoes_acesso", "observacoes_analise"}


def extrai(row):
    """Extrai e normaliza registro."""
    out = {}
    for campo, col in MAP.items():
        val = row.get(col, "")
        out[campo] = limpa_texto(val) if campo in TEXTO_LIMPO else (val or "").strip()

    # pesquisador (nomes diferentes entre forms)
    pesq = row.get("Pesquisador/a responsável pelo registro") or row.get(
        "Nome do estudante pesquisador que preencheu o formulário"
    )
    out["pesquisador"] = (pesq or "").strip()
    out["curso_fonte"] = (row.get("Curso/licenciatura analisada", "") or "").strip()
    out["grupo_tcc"] = grupo_canonico(out["curso_fonte"])

    return out


# ─────────────────────────────────────────────────────────────────────────────
# INTEGRAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
def integra(arquivo_novo):
    """Integra novos TCCs ao corpus."""
    arquivo = Path(arquivo_novo)
    if not arquivo.exists():
        print(f"❌ Arquivo não encontrado: {arquivo}")
        sys.exit(1)

    # carrega novo
    print(f"▶ Carregando novo arquivo: {arquivo}")
    with open(arquivo, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        novos_brutos = list(reader)

    print(f"✓ {len(novos_brutos)} registros no arquivo novo")

    # carrega base existente
    print(f"▶ Carregando base consolidada: {CONSOLIDADO_CSV}")
    base = pd.read_csv(CONSOLIDADO_CSV)
    print(f"✓ {len(base)} registros na base atual")

    # extrai e normaliza novos
    novos_normalizados = []
    for row in novos_brutos:
        rec = extrai(row)
        rec["em_ambos_forms"] = "Não"
        rec["conflitos"] = ""
        rec["fonte"] = "Integração nova"
        novos_normalizados.append(rec)

    # detecta duplicatas (por título normalizado)
    titulos_existentes = set(chave(t) for t in base["titulo"])
    duplicatas = 0
    novos_unicos = []

    for rec in novos_normalizados:
        t = chave(rec["titulo"])
        if t in titulos_existentes:
            duplicatas += 1
            print(f"⚠️  Duplicata detectada: {rec['titulo'][:60]}")
        else:
            novos_unicos.append(rec)
            titulos_existentes.add(t)

    print(f"\n✓ {len(novos_unicos)} registros novos (únicos)")
    print(f"⚠️  {duplicatas} duplicatas ignoradas")

    if not novos_unicos:
        print("\n❌ Nenhum registro novo para integrar!")
        return

    # adiciona novos à base
    df_novos = pd.DataFrame(novos_unicos)

    # ajusta IDs
    max_id = base["id"].max()
    df_novos["id"] = range(int(max_id) + 1, int(max_id) + len(df_novos) + 1)

    # reordena colunas para bater com a base
    cols_base = list(base.columns)
    cols_novos = [c for c in cols_base if c in df_novos.columns]
    for col in cols_base:
        if col not in df_novos.columns:
            df_novos[col] = ""

    df_novos = df_novos[cols_base]

    # concatena
    base_atualizada = pd.concat([base, df_novos], ignore_index=False)
    base_atualizada = base_atualizada.reset_index(drop=True)

    # backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_csv = BACKUP_DIR / f"corpus_tccs_analisado_BACKUP_{timestamp}.csv"
    base.to_csv(backup_csv, index=False, encoding="utf-8")
    print(f"\n✓ Backup criado: {backup_csv}")

    # salva
    base_atualizada.to_csv(CONSOLIDADO_CSV, index=False, encoding="utf-8")
    print(f"✓ Base atualizada: {CONSOLIDADO_CSV}")

    # regenera XLSX
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Corpus TCCs"

    ws.append(cols_base)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for _, row in base_atualizada.iterrows():
        ws.append([row[c] for c in cols_base])

    larguras = {"titulo": 50, "resumo": 80, "palavras_chave": 40,
                "autor": 28, "orientador": 28, "conflitos": 40}
    for idx, c in enumerate(cols_base, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = larguras.get(c, 16)

    ws.freeze_panes = "A2"
    wb.save(CONSOLIDADO_XLSX)
    print(f"✓ Excel atualizado: {CONSOLIDADO_XLSX}")

    # relatório
    print("\n" + "="*70)
    print("RESUMO DA INTEGRAÇÃO")
    print("="*70)
    print(f"\nRegistros adicionados: {len(df_novos)}")
    print(f"Total na base agora: {len(base_atualizada)}")
    print(f"\nNovos por grupo:")
    for g in sorted(df_novos["grupo_tcc"].unique()):
        count = (df_novos["grupo_tcc"] == g).sum()
        print(f"  - {g}: {count}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\n❌ USO: python3 integra_novos_tccs.py <arquivo_novo.csv>")
        sys.exit(1)

    arquivo = sys.argv[1]

    print("┌" + "─" * 68 + "┐")
    print("│ INTEGRAÇÃO DE NOVOS TCCs - CORPUS LIDAE/UFRR                    │")
    print("└" + "─" * 68 + "┘\n")

    integra(arquivo)

    print("\n✅ INTEGRAÇÃO CONCLUÍDA COM SUCESSO!")


if __name__ == "__main__":
    main()
