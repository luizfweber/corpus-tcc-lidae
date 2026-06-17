#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consolidação do corpus de TCCs das Licenciaturas UFRR (LIDAE) — versão "do zero".

Reconstrói a base de TCCs a partir das DUAS fontes originais (respostas dos
formulários Google), seguindo o CLAUDE.md:
  - Nunca imputar dados ausentes (marca e preserva "Não informado").
  - Preservar discrepâncias da fonte; nunca corrigir em silêncio.
  - Nomes canônicos de curso (grupo_tcc).
  - Sempre declarar como cada número foi obtido.

Saídas (corpus_v2/outputs/):
  - corpus_tccs_consolidado.csv   (base limpa, 1 linha por TCC único)
  - corpus_tccs_consolidado.xlsx  (mesma base em Excel)
  - relatorio_consolidacao.txt    (log do que foi feito / discrepâncias)
"""
import csv
import os
import re
import unicodedata
from collections import Counter, OrderedDict

# --------------------------------------------------------------------------- #
# Caminhos
# --------------------------------------------------------------------------- #
BASE_DADOS = (
    "/Users/lfernandojr/Library/CloudStorage/GoogleDrive-lfernandojr@gmail.com/"
    "Meu Drive/_UFRR_/[Grupo de Pesquisa]/[Projeto de pesquisa] UFRR - TCCs/"
    "Concatenar os dados/Dados até 12_junho_2026/"
)
FONTE_1 = BASE_DADOS + (
    "Catalogação de TCCs – Licenciaturas UFRR (respostas) - "
    "Respostas ao formulário 1 (1).csv"
)
FONTE_2 = BASE_DADOS + (
    "Catalogação de TCCs – Licenciaturas UFRR _ Projeto PROSUL_CNPq (respostas) - "
    "Respostas ao formulário 1 (1).csv"
)
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")

# --------------------------------------------------------------------------- #
# Normalização
# --------------------------------------------------------------------------- #
def chave(s):
    """Chave de agrupamento: sem acento, maiúscula, espaços colapsados."""
    s = (s or "").upper().strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s)


def limpa_texto(s):
    """Remove quebras de linha internas e colapsa espaços (preserva conteúdo)."""
    s = (s or "").replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip()


# Mapa curso (rótulo da fonte) -> grupo_tcc canônico (CLAUDE.md §5)
def grupo_canonico(curso):
    k = chave(curso)
    if k.startswith("PEDAGOGIA"):
        return "Pedagogia"
    if k.startswith("MATEMATICA"):
        return "Matemática"
    if k.startswith("MUSICA"):
        return "Música"
    if k.startswith("INSIKIRAN"):
        return "Insikiran"
    if k.startswith("LEDUCAR"):
        return "LEDUCAR"
    if k.startswith("LETRAS"):
        return "Letras"
    if k == "OUTRO" or k == "":
        return "Outro (a classificar)"
    return curso.strip()  # preserva grafia desconhecida em vez de inventar


# --------------------------------------------------------------------------- #
# Reclassificação manual (decisão humana, NÃO imputação automática).
# Registros catalogados como "Outro" na fonte, reclassificados por decisão do
# coordenador em 2026-06-16. A origem ("Outro") é preservada na coluna
# 'conflitos' para rastreabilidade (CLAUDE.md §3 e §5).
# Chave = título normalizado por chave().
# --------------------------------------------------------------------------- #
RECLASSIFICACAO_MANUAL = {
    "ALGORITMOS ALTERNATIVOS PARA AS OPERACOES ARITMETICAS FUNDAMENTAIS":
        ("Insikiran", "Insikiran – Ciências da Natureza"),
    "ALIMENTACAO, SAUDE E EDUCACAO NA COMUNIDADEINDIGENAMURIRU – TERRA INDIGENA MURIRU EMRORAIMA.":
        ("Insikiran", "Insikiran – Ciências da Natureza"),
    "APLICACAO DA MODELAGEM MATEMATICA NA CONSTRUCAO DE UMA HORTA ESCOLAR NA ESCOLA ESTADUAL INDIGENA NOVA MONTE MORIA II":
        ("Insikiran", "Insikiran – Ciências da Natureza"),
}


# --------------------------------------------------------------------------- #
# Leitura
# --------------------------------------------------------------------------- #
def carrega(path, fonte_tag):
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        r["_fonte"] = fonte_tag
    return rows


# Colunas canônicas de saída (subconjunto comum + curadoria)
COLS_SAIDA = [
    "id",
    "grupo_tcc",
    "curso_fonte",
    "titulo",
    "tipo_tcc",
    "autor",
    "orientador",
    "coorientador",
    "banca_examinadora",
    "ano_defesa",
    "semestre_defesa",
    "paginas",
    "palavras_chave",
    "resumo_disponivel",
    "resumo",
    "tcc_digital",
    "formato",
    "disponivel_integral",
    "link",
    "arquivo",
    "pesquisador",
    "observacoes_acesso",
    "observacoes_analise",
    "carimbo",
    "em_ambos_forms",
    "fonte",
    "conflitos",
]

# Mapeamento coluna-fonte -> campo-saída
# (pesquisador é tratado à parte: o nome da coluna difere entre os forms)
MAP = {
    "titulo": "Título completo do TCC",
    "tipo_tcc": "Tipo do TCC",  # só Form1
    "autor": "Autor/a do TCC",
    "orientador": "Orientador/a",
    "coorientador": "Coorientador/a, se houver",
    "banca_examinadora": "Membros da banca examinadora, se houver",
    "ano_defesa": "Ano de defesa",
    "semestre_defesa": "Semestre de defesa, se houver",
    "paginas": "Número de páginas",
    "palavras_chave": "Palavras-chave informadas no TCC",
    "resumo_disponivel": "Resumo disponível?",
    "resumo": "Resumo do TCC",
    "tcc_digital": "TCC digital?",
    "formato": "Formato do documento analisado",
    "disponivel_integral": "TCC disponível integralmente?",
    "link": "Link ou localização do TCC",
    "arquivo": "Envio do arquivo do TCC",
    "observacoes_acesso": "Observações sobre acesso ao documento",
    "observacoes_analise": "Observações da análise",  # só Form1
    "carimbo": "Carimbo de data/hora",
}

# campos onde removo \n (texto longo / multilinha na origem)
TEXTO_LIMPO = {"resumo", "titulo", "palavras_chave", "banca_examinadora",
               "observacoes_acesso", "observacoes_analise"}


def extrai(row):
    """Converte uma linha-fonte no dicionário de saída (sem id/dedup)."""
    out = {}
    for campo, col in MAP.items():
        val = row.get(col, "")
        out[campo] = limpa_texto(val) if campo in TEXTO_LIMPO else (val or "").strip()
    # pesquisador: nome da coluna difere entre os dois forms
    pesq = row.get("Pesquisador/a responsável pelo registro") or row.get(
        "Nome do estudante pesquisador que preencheu o formulário"
    )
    out["pesquisador"] = (pesq or "").strip()
    out["curso_fonte"] = (row.get("Curso/licenciatura analisada", "") or "").strip()
    out["grupo_tcc"] = grupo_canonico(out["curso_fonte"])
    out["fonte"] = row["_fonte"]
    return out


def detecta_conflitos(a, b):
    """Compara dois registros do mesmo TCC; retorna lista de conflitos legíveis."""
    confs = []
    for campo in ["ano_defesa", "paginas", "autor", "orientador"]:
        va, vb = (a.get(campo) or "").strip(), (b.get(campo) or "").strip()
        if chave(va) != chave(vb) and va and vb:
            confs.append(f"{campo}: F1='{va}' vs F2='{vb}'")
    return confs


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    log = []

    r1 = carrega(FONTE_1, "Form1")
    r2 = carrega(FONTE_2, "Form2")
    log.append(f"Fonte 1 (Licenciaturas UFRR):      {len(r1)} registros")
    log.append(f"Fonte 2 (PROSUL/CNPq):             {len(r2)} registros")
    log.append(f"Total bruto:                       {len(r1) + len(r2)} registros")

    # Indexa por chave de título
    consolidado = OrderedDict()  # chave_titulo -> registro de saída
    dups = 0

    # Form1 primeiro (mais rico em colunas) -> tem prioridade na dedup
    for row in r1:
        rec = extrai(row)
        k = chave(rec["titulo"])
        if not k:
            continue
        rec["em_ambos_forms"] = "Não"
        rec["conflitos"] = ""
        consolidado[k] = rec

    # Form2: se título já existe, é duplicata -> mantém Form1, registra conflito
    for row in r2:
        rec = extrai(row)
        k = chave(rec["titulo"])
        if not k:
            continue
        if k in consolidado:
            dups += 1
            base = consolidado[k]
            confs = detecta_conflitos(base, rec)
            base["em_ambos_forms"] = "Sim"
            base["fonte"] = "Form1+Form2"
            base["conflitos"] = " ; ".join(confs)
            # completa campos vazios do Form1 com Form2 (sem sobrescrever)
            for campo in MAP:
                if not (base.get(campo) or "").strip() and (rec.get(campo) or "").strip():
                    base[campo] = rec[campo]
        else:
            rec["em_ambos_forms"] = "Não"
            rec["conflitos"] = ""
            consolidado[k] = rec

    # Registros sem título real (ex.: "Monografia", "Artigo científico") são mantidos?
    # -> Aqui já foram incluídos se tinham qualquer título; nenhum foi descartado.

    # Numera
    final = list(consolidado.values())
    for i, rec in enumerate(final, 1):
        rec["id"] = i
        for c in COLS_SAIDA:
            rec.setdefault(c, "")

    # Reclassificação manual dos registros "Outro" (decisão humana documentada)
    reclass = 0
    for rec in final:
        kt = chave(rec["titulo"])
        if kt in RECLASSIFICACAO_MANUAL:
            novo_grupo, novo_curso = RECLASSIFICACAO_MANUAL[kt]
            origem = rec["curso_fonte"]
            nota = (f"reclassificado manualmente de '{origem}' para "
                    f"'{novo_curso}' (decisão do coordenador, 2026-06-16)")
            rec["grupo_tcc"] = novo_grupo
            rec["curso_fonte"] = novo_curso
            rec["conflitos"] = (rec["conflitos"] + " ; " + nota).strip(" ;")
            reclass += 1
    nao_casou = len(RECLASSIFICACAO_MANUAL) - reclass
    if nao_casou:
        log.append(f"AVISO: {nao_casou} título(s) de reclassificação não casaram!")

    log.append(f"Duplicatas (título em ambos forms): {dups}")
    log.append(f"TCCs únicos consolidados:          {len(final)}")
    log.append(f"Reclassificações manuais (Outro→):  {reclass}")
    log.append("")

    # Distribuição por grupo
    log.append("Distribuição por grupo_tcc (TCCs coletados):")
    dist = Counter(r["grupo_tcc"] for r in final)
    for g, n in dist.most_common():
        log.append(f"  {g:<24} {n}")
    log.append("")

    # Resumos: quantos disponíveis / com texto
    com_resumo = sum(1 for r in final if len(r["resumo"]) > 20)
    log.append(f"Registros com resumo (>20 chars):  {com_resumo}/{len(final)}")

    def preenchido(v):
        return str(v or "").strip() not in ("", "Não informado", "Não se aplica", "Não")

    com_banca = sum(1 for r in final if preenchido(r["banca_examinadora"]))
    log.append(f"Registros com banca examinadora:   {com_banca}/{len(final)} "
                "(campo vindo das fontes originais; não imputado)")

    # Conflitos preservados
    confl = [r for r in final if r["conflitos"]]
    log.append(f"Registros com conflito de campo:   {len(confl)} (preservados na coluna 'conflitos')")

    # ---- grava CSV ----
    csv_path = os.path.join(OUT_DIR, "corpus_tccs_consolidado.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS_SAIDA)
        w.writeheader()
        for r in final:
            w.writerow({c: r.get(c, "") for c in COLS_SAIDA})

    # ---- grava XLSX ----
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Corpus TCCs"
    ws.append(COLS_SAIDA)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for r in final:
        ws.append([r.get(c, "") for c in COLS_SAIDA])
    # larguras razoáveis
    larguras = {"titulo": 50, "resumo": 80, "palavras_chave": 40, "autor": 28,
                "orientador": 28, "conflitos": 40}
    for idx, c in enumerate(COLS_SAIDA, 1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col_letter].width = larguras.get(c, 16)
    ws.freeze_panes = "A2"
    xlsx_path = os.path.join(OUT_DIR, "corpus_tccs_consolidado.xlsx")
    wb.save(xlsx_path)

    # ---- grava log ----
    rel = os.path.join(OUT_DIR, "relatorio_consolidacao.txt")
    with open(rel, "w", encoding="utf-8") as f:
        f.write("RELATÓRIO DE CONSOLIDAÇÃO DO CORPUS DE TCCs — LIDAE/UFRR\n")
        f.write("=" * 60 + "\n\n")
        f.write("\n".join(log) + "\n\n")
        if confl:
            f.write("DETALHE DOS CONFLITOS PRESERVADOS:\n")
            for r in confl:
                f.write(f"  [id {r['id']}] {r['titulo'][:60]}\n      {r['conflitos']}\n")

    print("\n".join(log))
    print(f"\nArquivos gerados em: {OUT_DIR}")
    print("  - corpus_tccs_consolidado.csv")
    print("  - corpus_tccs_consolidado.xlsx")
    print("  - relatorio_consolidacao.txt")


if __name__ == "__main__":
    main()
