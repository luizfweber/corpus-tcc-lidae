#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reimporta dados preenchidos de tipo_tcc para a base consolidada.

USO:
  python3 reimporta_tipo_tcc.py <arquivo_preenchido>

EXEMPLO:
  python3 reimporta_tipo_tcc.py outputs/relatorio_tipo_tcc_vazio_PREENCHIDO.xlsx

ARQUIVO ESPERADO:
  - Coluna A (id): ID do TCC
  - Coluna H (tipo_tcc): tipo preenchido (Monografia, Artigo, etc.)
  - Formatos aceitos: .xlsx, .csv

VALIDAÇÕES:
  - Verifica se id existe na base
  - Verifica se tipo_tcc não é vazio
  - Registra conflitos (ex.: TCC já tinha tipo preenchido)
  - Cria backup antes de atualizar
"""
import sys
from pathlib import Path
import pandas as pd
import csv
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
BASE_CSV = BASE_DIR / "outputs" / "analise" / "corpus_tccs_analisado.csv"
BASE_XLSX = BASE_DIR / "outputs" / "analise" / "corpus_tccs_analisado.xlsx"
BACKUP_DIR = BASE_DIR / "outputs" / "backups"
BACKUP_DIR.mkdir(exist_ok=True)

TIPOS_VALIDOS = {
    "monografia", "artigo", "relato de experiência",
    "trabalho de pesquisa", "projeto", "dissertação",
    "tese", "estudo de caso", "revisão sistemática",
    "ensaio", "resenha", "outro"
}


def carrega_preenchido(arquivo):
    """Carrega o arquivo preenchido (XLSX ou CSV)."""
    arquivo = Path(arquivo)
    if not arquivo.exists():
        print(f"❌ Arquivo não encontrado: {arquivo}")
        sys.exit(1)

    if arquivo.suffix == ".xlsx":
        df = pd.read_excel(arquivo)
    elif arquivo.suffix == ".csv":
        df = pd.read_csv(arquivo)
    else:
        print(f"❌ Formato não suportado: {arquivo.suffix}")
        sys.exit(1)

    print(f"✓ Carregado: {len(df)} registros")
    return df


def valida_dados(df_novo):
    """Valida os dados antes de importar."""
    erros = []

    # verifica coluna id
    if "id" not in df_novo.columns:
        erros.append("Coluna 'id' não encontrada")
    else:
        ids_vazio = df_novo[df_novo["id"].isna() | (df_novo["id"].astype(str).str.strip() == "")].shape[0]
        if ids_vazio > 0:
            erros.append(f"  ⚠️  {ids_vazio} registros com id vazio (serão pulados)")

    # verifica coluna tipo_tcc
    if "tipo_tcc" not in df_novo.columns:
        erros.append("Coluna 'tipo_tcc' não encontrada")
    else:
        vazios = df_novo[df_novo["tipo_tcc"].isna() | (df_novo["tipo_tcc"].astype(str).str.strip() == "")].shape[0]
        if vazios > 0:
            print(f"⚠️  {vazios} registros com tipo_tcc ainda vazio (serão ignorados)")
        # tipos inválidos
        tipos_unicos = df_novo["tipo_tcc"].dropna().astype(str).str.strip().str.lower().unique()
        invalidos = [t for t in tipos_unicos if t and t not in TIPOS_VALIDOS]
        if invalidos:
            print(f"⚠️  Tipos não reconhecidos (podem estar corretos): {invalidos}")

    if erros:
        print("❌ ERROS DE VALIDAÇÃO:")
        for e in erros:
            print(f"  {e}")
        return False
    return True


def reimporta(df_novo):
    """Reimporta dados para a base consolidada."""
    # carrega base original
    df_base = pd.read_csv(BASE_CSV)
    print(f"✓ Base original carregada: {len(df_base)} registros")

    # backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_csv = BACKUP_DIR / f"corpus_tccs_analisado_BACKUP_{timestamp}.csv"
    df_base.to_csv(backup_csv, index=False, encoding="utf-8")
    print(f"✓ Backup criado: {backup_csv}")

    # atualiza
    atualizados = 0
    conflitos = []
    nao_encontrados = []

    for _, row in df_novo.iterrows():
        id_val = row.get("id")
        tipo_novo = str(row.get("tipo_tcc", "")).strip()

        # valida
        if not id_val or pd.isna(id_val):
            continue
        if not tipo_novo:
            continue

        id_val = int(id_val)

        # procura na base
        mask = df_base["id"] == id_val
        if not mask.any():
            nao_encontrados.append(id_val)
            continue

        # verifica conflito (tipo já preenchido)
        tipo_antigo = df_base.loc[mask, "tipo_tcc"].values[0]
        if tipo_antigo and str(tipo_antigo).strip():
            conflitos.append({
                "id": id_val,
                "tipo_antigo": tipo_antigo,
                "tipo_novo": tipo_novo,
                "acao": "preservado o antigo"
            })
            continue

        # atualiza
        df_base.loc[mask, "tipo_tcc"] = tipo_novo
        atualizados += 1

    # salva
    df_base.to_csv(BASE_CSV, index=False, encoding="utf-8")
    print(f"✓ Base atualizada: {BASE_CSV}")

    # atualiza XLSX também
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Corpus TCCs"

    cols = list(df_base.columns)
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for _, row in df_base.iterrows():
        ws.append([row[c] for c in cols])

    larguras = {
        "titulo": 50, "resumo": 80, "palavras_chave": 40,
        "autor": 28, "orientador": 28, "conflitos": 40
    }
    for idx, c in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = larguras.get(c, 16)

    ws.freeze_panes = "A2"
    wb.save(BASE_XLSX)
    print(f"✓ Excel atualizado: {BASE_XLSX}")

    # relatório
    print("\n" + "="*70)
    print("RELATÓRIO DE REIMPORTAÇÃO")
    print("="*70)
    print(f"\n✓ Registros atualizados: {atualizados}")

    if conflitos:
        print(f"\n⚠️  Conflitos encontrados: {len(conflitos)}")
        print("   (tipo_tcc já estava preenchido — mantido valor original)")
        for conf in conflitos[:5]:  # mostra primeiros 5
            print(f"     id {conf['id']}: {conf['tipo_antigo']} (novo: {conf['tipo_novo']})")
        if len(conflitos) > 5:
            print(f"     ... e mais {len(conflitos) - 5}")

    if nao_encontrados:
        print(f"\n❌ IDs não encontrados na base: {len(nao_encontrados)}")
        print(f"   {nao_encontrados[:10]}")

    print(f"\n📊 Resumo final:")
    tipos_count = df_base[df_base["tipo_tcc"].notna() & (df_base["tipo_tcc"].astype(str).str.strip() != "")]["tipo_tcc"].value_counts()
    print(f"   Total com tipo_tcc preenchido: {(df_base['tipo_tcc'].notna()).sum()}")
    print(f"   Distribuição:")
    for tipo, count in tipos_count.items():
        print(f"     - {tipo}: {count}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\n❌ USO: python3 reimporta_tipo_tcc.py <arquivo_preenchido.xlsx ou .csv>")
        sys.exit(1)

    arquivo = sys.argv[1]

    print("┌" + "─" * 68 + "┐")
    print("│ REIMPORTAÇÃO DE TIPO_TCC - CORPUS LIDAE/UFRR                     │")
    print("└" + "─" * 68 + "┘\n")

    # carrega
    df_novo = carrega_preenchido(arquivo)

    # valida
    print("\n▶ Validando dados...")
    if not valida_dados(df_novo):
        sys.exit(1)

    # reimporta
    print("\n▶ Reimportando...")
    reimporta(df_novo)

    print("\n✅ CONCLUÍDO COM SUCESSO!")
    print(f"\nBackup salvo em: {BACKUP_DIR}")


if __name__ == "__main__":
    main()
